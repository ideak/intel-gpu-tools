#!/usr/bin/env python3
# pylint: disable=C0103, C0200
# SPDX-License-Identifier: (GPL-2.0 OR MIT)

## Copyright (C) 2023    Intel Corporation                 ##
## Author: Mauro Carvalho Chehab <mchehab@kernel.org>      ##
##                                                         ##
## Allow keeping inlined test documentation and validate   ##
## if the documentation is kept updated.                   ##

"""Write the contents of the testplan documentation to a XLS file."""

EPILOG="""
Examples:

1. Create a XLS file with a single worksheet with Xe driver documentation:

   scripts/doc_to_xls.py --config tests/kms_*json tests/*/*.json --xls igt_test_documentation.xls

2. Create a XLS file with one sheet per driver, for all drivers with testplan config files and KMS:

   scripts/doc_to_xls.py --config tests/kms_*json tests/*/*.json --xls igt_test_documentation.xls
"""

import argparse

from openpyxl import Workbook

from test_list import TestList

parser = argparse.ArgumentParser(description=__doc__,
                                 formatter_class = argparse.RawDescriptionHelpFormatter,
                                 epilog = EPILOG)
parser.add_argument("--config", required = True,  nargs='+',
                    help="JSON file describing the test plan template")
parser.add_argument("--include-plan", action="store_true",
                    help="Include test plans, if any.")
parser.add_argument("--xls", required = True,
                    help="Output XLS file.")

parse_args = parser.parse_args()

tests = []
for config_file in parse_args.config:
    # Implemented tests
    tests.append(TestList(config_file, parse_args.include_plan))

wb = Workbook()
ws = None

for row in range(len(tests)):
    test = tests[row]
    sheet_name = test.title

    if not ws:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    sheet = test.get_spreadsheet()

    max_length = []
    for col in range(len(sheet[row])):
        max_length.append(0)

    for row in range(len(sheet)):
        for col in range(len(sheet[row])):
            ws.cell(row = row + 1, column = col + 1, value = sheet[row][col])
            if len(sheet[row][col]) > max_length[col]:
                max_length[col] = len(sheet[row][col])

    col = 0
    for c in ws.columns:
        column = c[0].column_letter

        adjusted_width = (max_length[col] + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

        col += 1

wb.save(parse_args.xls)
